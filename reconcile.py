"""
Reconcile Reliance portal export vs internal billing annexure (TCN-driven),
with Vendor Summary CSV as fallback for empty Reliance fields.

Flow per BILLING row (Annexure):
  1. Take TCN/Trip Number from Billing.
  2. Look it up in Reliance sheet by TCN/Trip Number.
     - Not found  -> NOT_IN_RELIANCE
  3. If found, apply Vendor Summary fallback to empty Reliance fields
     (looked up via the same TCN/Trip Number).
  4. Check Tms Status.
     - Rejected   -> REJECTED
     - else       -> compare all mapped fields
  5. Any diff -> MISMATCH, else MATCHED.
"""
from openpyxl import load_workbook
from io import BytesIO
from pathlib import Path
import csv
import re
import json


# -------- loaders --------

def fix_corrupted_xlsx(data: bytes) -> bytes:
    """Trivedi billing portal appends HTML to xlsx; truncate at ZIP EOCD."""
    eocd = data.rfind(b'PK\x05\x06')
    if eocd < 0:
        return data
    comment_len = int.from_bytes(data[eocd + 20:eocd + 22], 'little')
    return data[:eocd + 22 + comment_len]


def _sheet_to_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    return [dict(zip(headers, r)) for r in rows[1:] if any(c is not None and c != '' for c in r)]


def load_billing_annexure(path_or_bytes, sheet='Annexure'):
    data = Path(path_or_bytes).read_bytes() if isinstance(path_or_bytes, str) else path_or_bytes
    wb = load_workbook(BytesIO(fix_corrupted_xlsx(data)), data_only=True)
    return _sheet_to_dicts(wb[sheet])


def load_reliance(path_or_bytes, sheet='Sheet1'):
    data = Path(path_or_bytes).read_bytes() if isinstance(path_or_bytes, str) else path_or_bytes
    wb = load_workbook(BytesIO(data), data_only=True)
    return _sheet_to_dicts(wb[sheet])


def load_vendor_summary(path_or_bytes):
    if isinstance(path_or_bytes, str):
        with open(path_or_bytes, newline='', encoding='utf-8') as f:
            return list(csv.DictReader(f))
    text = path_or_bytes.decode('utf-8') if isinstance(path_or_bytes, bytes) else path_or_bytes
    return list(csv.DictReader(text.splitlines()))


# -------- helpers --------

def is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == '')


def norm_str(v):
    return '' if v is None else str(v).strip().upper()


def norm_num(v):
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(',', ''))
    except ValueError:
        return None


def parse_mt_to_kg(v):
    """Parses '9.000 MT' or '9MT 22FT Open' -> 9000 kg."""
    if v is None:
        return None
    m = re.search(r'([\d.]+)\s*MT', str(v), re.IGNORECASE)
    return float(m.group(1)) * 1000 if m else None


# -------- vendor-summary fallback --------

VENDOR_FALLBACK = {
    'SO Number/Ref No.': 'Planned Consignment',
    'Vehicle Number':    'Vehicle Number',
    'Freight Cost':      'Freight Cost',
    'Weight':            'Vehicle Make Name',
}


def build_vendor_index(vendor_rows):
    """Trip Number -> first row (duplicates are same-trip multi-SO; trip-level fields equal)."""
    idx = {}
    for row in vendor_rows:
        tn = norm_str(row.get('Trip Number'))
        if tn and tn not in idx:
            idx[tn] = row
    return idx


def apply_vendor_fallback(rel_row, vendor_index):
    filled = dict(rel_row)
    used = []
    tn = norm_str(rel_row.get('TCN/Trip Number'))
    if not tn or tn not in vendor_index:
        return filled, used
    v = vendor_index[tn]
    for rel_col, vendor_col in VENDOR_FALLBACK.items():
        if is_empty(filled.get(rel_col)):
            val = v.get(vendor_col)
            if is_empty(val):
                continue
            if rel_col == 'Weight':
                kg = parse_mt_to_kg(val)
                if kg is not None:
                    filled[rel_col] = kg
                    used.append(rel_col)
            else:
                filled[rel_col] = val
                used.append(rel_col)
    return filled, used


# -------- field comparison map (Reliance col -> Billing col) --------

FIELD_MAP = [
    ('SO Number/Ref No.',   'SO Number',                        'str'),
    ('Vehicle Number',      'Vehicle Number',                   'str'),
    ('Weight',              'Vehicle Type',                     'weight'),
    ('Freight Cost',        'Frieght as per Emptoris Contract', 'num'),
    ('Loading Charges',     'Load Charges',                     'num'),
    ('Unloading Charges',   'Unload charges',                   'num'),
    ('Other Charges',       'Other charges',                    'num'),
    ('Detentation Charges', 'Detention Charges',                'num'),
    ('Invoice No.',         'Invoice Number',                   'str'),
    ('Internal Ref No.',    'Internal Ref No',                  'str'),
]


def compare_row(rel, bill):
    """Returns (matched, mismatched) where each is a list of dicts with
    field name + reliance value + billing value."""
    matched, mismatched = [], []
    for r_col, b_col, kind in FIELD_MAP:
        r_val, b_val = rel.get(r_col), bill.get(b_col)
        if kind == 'str':
            ok = norm_str(r_val) == norm_str(b_val)
        elif kind == 'num':
            rn, bn = norm_num(r_val), norm_num(b_val)
            ok = rn is not None and bn is not None and abs(rn - bn) < 0.01
        elif kind == 'weight':
            rn = norm_num(r_val)
            bn = parse_mt_to_kg(b_val) if not isinstance(b_val, (int, float)) else float(b_val)
            ok = rn is not None and bn is not None and abs(rn - bn) < 0.01
        else:
            ok = False
        item = {'field': r_col, 'reliance': r_val, 'billing': b_val}
        (matched if ok else mismatched).append(item)
    return matched, mismatched


# -------- main reconcile (BILLING-driven, TCN match) --------

def reconcile(rel_rows, bill_rows, vendor_rows=None):
    # Index Reliance by TCN/Trip Number (TCN is unique)
    rel_index = {}
    for row in rel_rows:
        tcn = norm_str(row.get('TCN/Trip Number'))
        if tcn:
            rel_index[tcn] = row

    vendor_index = build_vendor_index(vendor_rows) if vendor_rows else {}

    results = []
    for b in bill_rows:
        tcn = norm_str(b.get('TCN/Trip Number'))
        entry = {
            'billing_so': b.get('SO Number'),
            'tcn_trip': b.get('TCN/Trip Number'),
            'billing_invoice': b.get('Invoice Number'),
            'billing_row': b,
        }

        # Stage 1: must exist in Reliance
        if not tcn or tcn not in rel_index:
            entry['result'] = 'NOT_IN_RELIANCE'
            entry['tms_status'] = None
            entry['fallback_fields'] = []
            entry['mismatches'] = []
            results.append(entry)
            continue

        rel_row = rel_index[tcn]
        filled, used = apply_vendor_fallback(rel_row, vendor_index)
        entry['reliance_row'] = rel_row
        entry['filled_row'] = filled
        entry['fallback_fields'] = used
        entry['tms_status'] = filled.get('Tms Status')

        # Stage 2: TMS status
        status = norm_str(filled.get('Tms Status'))
        if status == 'REJECTED':
            entry['result'] = 'REJECTED'
            entry['rcpl_remark'] = filled.get('RCPL Remark')
            entry['mismatches'] = []
            results.append(entry)
            continue

        # Stage 3: field-by-field
        matched, mismatches = compare_row(filled, b)
        entry['matched_fields'] = matched
        entry['mismatches'] = mismatches
        entry['result'] = 'MATCHED' if not mismatches else 'MISMATCH'
        results.append(entry)

    return results


def print_report(results):
    counts = {}
    for e in results:
        counts[e['result']] = counts.get(e['result'], 0) + 1
        fb = f" [fallback: {','.join(e['fallback_fields'])}]" if e['fallback_fields'] else ''
        print(f"TCN {str(e['tcn_trip']):<14} | SO {str(e['billing_so']):<12} | TMS {str(e['tms_status']):<10} | {e['result']}{fb}")
        for m in e.get('mismatches', []):
            print(f"     ✗ {m['field']:<22} reliance={m['reliance']!r}  billing={m['billing']!r}")
        for m in e.get('matched_fields', []):
            print(f"     ✓ {m['field']:<22} {m['reliance']!r}")
    print("\n--- Summary ---")
    for k, v in counts.items():
        print(f"  {k}: {v}")
    print(f"  TOTAL: {len(results)}")


if __name__ == '__main__':
    import sys
    rel = sys.argv[1] if len(sys.argv) > 1 else 'reliance.xlsx'
    bill = sys.argv[2] if len(sys.argv) > 2 else 'billing.xlsx'
    vendor = sys.argv[3] if len(sys.argv) > 3 else None

    rel_rows = load_reliance(rel)
    bill_rows = load_billing_annexure(bill)
    vendor_rows = load_vendor_summary(vendor) if vendor else None

    out = reconcile(rel_rows, bill_rows, vendor_rows)
    print_report(out)

    Path('reconciliation_result.json').write_text(json.dumps(out, default=str, indent=2))
    print("\nSaved: reconciliation_result.json")