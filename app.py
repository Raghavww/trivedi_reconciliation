"""
Streamlit UI for Trivedi <-> Reliance reconciliation (TCN-driven).
Run: streamlit run app.py
"""
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO, StringIO
import csv
import re

st.set_page_config(page_title="Trivedi Reconciliation", layout="wide", page_icon="🚛")

# ─── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Fonts & base ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* ── Page background ── */
.stApp { background: #F7F8FA; }

/* ── Header banner ── */
.header-banner {
    background: linear-gradient(135deg, #1B2A4A 0%, #2C4A8A 100%);
    border-radius: 16px;
    padding: 28px 36px;
    margin-bottom: 28px;
    display: flex;
    align-items: center;
    gap: 18px;
}
.header-icon { font-size: 42px; }
.header-title { color: #fff; font-size: 26px; font-weight: 700; margin: 0; }
.header-sub   { color: #A8BFE8; font-size: 13px; margin: 4px 0 0; }

/* ── Upload card ── */
.upload-card {
    background: #fff;
    border: 2px dashed #CBD5E1;
    border-radius: 14px;
    padding: 20px 16px 14px;
    text-align: center;
    transition: border-color .2s;
}
.upload-card:hover { border-color: #2C4A8A; }
.upload-label {
    font-size: 12px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: .06em;
    color: #64748B;
    margin-bottom: 8px;
}
.upload-badge {
    display: inline-block;
    font-size: 10px;
    background: #EFF6FF;
    color: #2C4A8A;
    border-radius: 20px;
    padding: 2px 10px;
    margin-top: 6px;
    font-weight: 500;
}

/* ── Metric cards ── */
.metric-row { display: flex; gap: 14px; margin: 20px 0 10px; }
.metric-card {
    flex: 1;
    border-radius: 14px;
    padding: 18px 20px;
    border: 1.5px solid transparent;
}
.metric-card.green  { background:#F0FDF4; border-color:#86EFAC; }
.metric-card.amber  { background:#FFFBEB; border-color:#FCD34D; }
.metric-card.red    { background:#FFF1F2; border-color:#FDA4AF; }
.metric-card.slate  { background:#F8FAFC; border-color:#CBD5E1; }
.metric-num  { font-size: 36px; font-weight: 700; line-height: 1; }
.metric-lbl  { font-size: 12px; font-weight: 600; margin-top: 4px; text-transform: uppercase; letter-spacing: .05em; }
.green  .metric-num { color: #16A34A; } .green  .metric-lbl { color: #15803D; }
.amber  .metric-num { color: #D97706; } .amber  .metric-lbl { color: #B45309; }
.red    .metric-num { color: #E11D48; } .red    .metric-lbl { color: #BE123C; }
.slate  .metric-num { color: #475569; } .slate  .metric-lbl { color: #64748B; }

/* ── Section heading ── */
.section-heading {
    font-size: 13px;
    font-weight: 700;
    color: #94A3B8;
    text-transform: uppercase;
    letter-spacing: .08em;
    margin: 24px 0 10px;
}

/* ── Reconcile button ── */
.stButton > button {
    background: linear-gradient(135deg, #1B2A4A, #2C4A8A) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    padding: 10px 32px !important;
    font-size: 15px !important;
    letter-spacing: .02em !important;
    transition: opacity .2s !important;
}
.stButton > button:hover { opacity: .88 !important; }

/* ── Dataframe ── */
.stDataFrame { border-radius: 12px; overflow: hidden; }

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: #fff;
    border-radius: 12px 12px 0 0;
    padding: 6px 12px 0;
    border-bottom: 2px solid #E2E8F0;
    gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px 8px 0 0 !important;
    font-weight: 600 !important;
    font-size: 13.5px !important;
    color: #64748B !important;
    padding: 8px 20px !important;
}
.stTabs [aria-selected="true"] {
    background: #EFF6FF !important;
    color: #1D4ED8 !important;
    border-bottom: 2px solid #1D4ED8 !important;
}

/* ── Search bar ── */
.stTextInput > div > div > input {
    border-radius: 10px !important;
    border: 1.5px solid #E2E8F0 !important;
    padding: 10px 14px !important;
    font-size: 14px !important;
}
.stTextInput > div > div > input:focus {
    border-color: #2C4A8A !important;
    box-shadow: 0 0 0 3px rgba(44,74,138,.12) !important;
}

/* ── Multiselect pills ── */
.stMultiSelect [data-baseweb="tag"] {
    background: #EFF6FF !important;
    color: #1D4ED8 !important;
    border-radius: 20px !important;
}

/* ── Download button ── */
.stDownloadButton > button {
    background: #fff !important;
    color: #1B2A4A !important;
    border: 1.5px solid #CBD5E1 !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 13px !important;
}

/* ── Empty state ── */
.empty-state {
    text-align: center;
    padding: 52px 20px;
    background: #fff;
    border-radius: 16px;
    border: 2px dashed #E2E8F0;
}
.empty-state-icon { font-size: 48px; margin-bottom: 12px; }
.empty-state-text { color: #64748B; font-size: 15px; font-weight: 500; }
.empty-state-sub  { color: #94A3B8; font-size: 13px; margin-top: 6px; }
</style>
""", unsafe_allow_html=True)


# ─── Loaders ───────────────────────────────────────────────────────────────────

def fix_corrupted_xlsx(data):
    eocd = data.rfind(b'PK\x05\x06')
    if eocd < 0: return data
    cl = int.from_bytes(data[eocd+20:eocd+22], 'little')
    return data[:eocd+22+cl]

def _sheet_to_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    return [dict(zip(headers, r)) for r in rows[1:] if any(c not in (None,'') for c in r)]

def load_billing(data, sheet='Annexure'):
    wb = load_workbook(BytesIO(fix_corrupted_xlsx(data)), data_only=True)
    return _sheet_to_dicts(wb[sheet])

def load_reliance(data, sheet='Sheet1'):
    wb = load_workbook(BytesIO(data), data_only=True)
    return _sheet_to_dicts(wb[sheet])

def load_vendor_summary(data):
    return list(csv.DictReader(StringIO(data.decode('utf-8'))))

def load_detention(data):
    wb = load_workbook(BytesIO(data), data_only=True)
    all_rows = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        hidx = next((i for i,r in enumerate(rows)
                     if any(c and 'lr' in str(c).lower() for c in r)), None)
        if hidx is None: continue
        headers = [str(h).strip().lower() if h else '' for h in rows[hidx]]
        lr_col  = next((i for i,h in enumerate(headers)
                        if h in ('lr n.','lr no','lr no.','lr n','lr number','lr n. ')), None)
        amt_col = next((i for i,h in enumerate(headers)
                        if any(k in h for k in ('unlod','unload','amount'))), None)
        if lr_col is None: continue
        for row in rows[hidx+1:]:
            lr_val  = row[lr_col]  if lr_col  < len(row) else None
            amt_val = row[amt_col] if amt_col is not None and amt_col < len(row) else None
            if not lr_val or str(lr_val).strip() == '': continue
            all_rows.append({'sheet': sname, 'lr_number': str(lr_val).strip(),
                             'detention_unload_amount': amt_val})
    return all_rows


# ─── Helpers ───────────────────────────────────────────────────────────────────

def is_empty(v): return v is None or (isinstance(v, str) and v.strip() == '')
def norm_str(v): return '' if v is None else str(v).strip().upper()

def norm_num(v):
    if v is None or v == '': return 0.0
    if isinstance(v, (int,float)): return float(v)
    try: return float(str(v).strip().replace(',',''))
    except: return None

def parse_mt_to_kg(v):
    if v is None: return None
    m = re.search(r'([\d.]+)\s*MT', str(v), re.IGNORECASE)
    return float(m.group(1))*1000 if m else None


# ─── Vendor fallback ───────────────────────────────────────────────────────────

VENDOR_FALLBACK = {
    'SO Number/Ref No.': 'Planned Consignment',
    'Vehicle Number':    'Vehicle Number',
    'Freight Cost':      'Freight Cost',
    'Weight':            'Vehicle Make Name',
}

def build_vendor_index(rows):
    idx = {}
    for r in rows:
        tn = norm_str(r.get('Trip Number'))
        if tn and tn not in idx: idx[tn] = r
    return idx

def apply_vendor_fallback(rel, vidx):
    filled, used = dict(rel), []
    tn = norm_str(rel.get('TCN/Trip Number'))
    if not tn or tn not in vidx: return filled, used
    v = vidx[tn]
    for rc, vc in VENDOR_FALLBACK.items():
        if is_empty(filled.get(rc)):
            val = v.get(vc)
            if is_empty(val): continue
            if rc == 'Weight':
                kg = parse_mt_to_kg(val)
                if kg: filled[rc] = kg; used.append(rc)
            else: filled[rc] = val; used.append(rc)
    return filled, used


# ─── Field map ─────────────────────────────────────────────────────────────────

FIELD_MAP = [
    ('SO Number/Ref No.',   'SO Number',                        'str'),
    ('Vehicle Number',      'Vehicle Number',                   'str'),
    ('Weight',              'Vehicle Type',                     'weight'),
    ('Freight Cost',        'Frieght as per Emptoris Contract', 'num'),
    ('Loading Charges',     'Load Charges',                     'num'),
    ('Unloading Charges',   'Unload charges',                   'num'),
    ('Other Charges',       'Other charges',                    'num'),
    ('Detentation Charges', 'Detention Charges',                'num'),
    ('Invoice No.',         'Invoice Number',                   'str'),
    ('Internal Ref No.',    'Internal Ref No',                  'str'),
]

def compare_row(rel, bill):
    matched, mismatched = [], []
    for rc, bc, kind in FIELD_MAP:
        rv, bv = rel.get(rc), bill.get(bc)
        if kind == 'str':   ok = norm_str(rv) == norm_str(bv)
        elif kind == 'num':
            rn,bn = norm_num(rv), norm_num(bv)
            ok = rn is not None and bn is not None and abs(rn-bn)<0.01
        else:
            rn = norm_num(rv)
            bn = parse_mt_to_kg(bv) if not isinstance(bv,(int,float)) else float(bv)
            ok = rn is not None and bn is not None and abs(rn-bn)<0.01
        (matched if ok else mismatched).append({'field':rc,'reliance':rv,'billing':bv})
    return matched, mismatched


# ─── Reconcile ─────────────────────────────────────────────────────────────────

def reconcile(rel_rows, bill_rows, vendor_rows):
    rel_idx = {norm_str(r.get('TCN/Trip Number')): r for r in rel_rows
               if norm_str(r.get('TCN/Trip Number'))}
    vidx = build_vendor_index(vendor_rows) if vendor_rows else {}
    results = []
    for b in bill_rows:
        tcn = norm_str(b.get('TCN/Trip Number'))
        entry = {'TCN/Trip': b.get('TCN/Trip Number'), 'SO Number': b.get('SO Number'),
                 'Billing Invoice': b.get('Invoice Number'),
                 'Tms Status':'', 'RCPL Remark':'', 'Fallback Used':'',
                 'Matched Fields':'', 'Mismatches':''}
        if not tcn or tcn not in rel_idx:
            entry['Result'] = '❌ NOT IN RELIANCE'; results.append(entry); continue
        filled, used = apply_vendor_fallback(rel_idx[tcn], vidx)
        entry['Tms Status']   = filled.get('Tms Status') or ''
        entry['RCPL Remark']  = filled.get('RCPL Remark') or ''
        entry['Fallback Used']= ', '.join(used) if used else ''
        if norm_str(filled.get('Tms Status')) == 'REJECTED':
            entry['Result'] = '🚫 REJECTED'; results.append(entry); continue
        matched, mismatches = compare_row(filled, b)
        entry['Matched Fields'] = ', '.join(m['field'] for m in matched)
        if not mismatches:
            entry['Result'] = '✅ MATCHED'
        else:
            entry['Result'] = '⚠️ MISMATCH'
            entry['Mismatches'] = ' | '.join(
                f"{m['field']}: R={m['reliance']} vs B={m['billing']}" for m in mismatches)
        results.append(entry)
    return results

def reconcile_detention(bill_rows, det_rows):
    det_idx = {}
    for r in det_rows:
        lr = norm_str(r.get('lr_number'))
        if lr: det_idx.setdefault(lr,[]).append(r)
    results = []
    for b in bill_rows:
        lr = norm_str(b.get('LR No.'))
        bu = norm_num(b.get('Unload charges'))
        entry = {'LR No': b.get('LR No.'), 'SO Number': b.get('SO Number'),
                 'TCN/Trip': b.get('TCN/Trip Number'),
                 'Billing Unload Amt': bu, 'Detention Unload Amt':'', 'Detention Sheet':''}
        if not lr or lr not in det_idx:
            entry['Result'] = '❌ LR NOT IN DETENTION'; entry['Diff']=''; results.append(entry); continue
        dr   = det_idx[lr][0]
        da   = norm_num(dr.get('detention_unload_amount'))
        entry['Detention Unload Amt'] = da
        entry['Detention Sheet']      = dr.get('sheet','')
        if bu is not None and da is not None and abs(bu-da)<0.01:
            entry['Result']='✅ MATCHED'; entry['Diff']=0
        else:
            entry['Result']='⚠️ MISMATCH'; entry['Diff']=(bu or 0)-(da or 0)
        results.append(entry)
    return results


# ─── UI helpers ────────────────────────────────────────────────────────────────

def metric_card(num, label, kind):
    st.markdown(f"""
    <div class="metric-card {kind}">
        <div class="metric-num">{num}</div>
        <div class="metric-lbl">{label}</div>
    </div>""", unsafe_allow_html=True)

def render_table(df, key_prefix):
    """Search bar + result filter + colored dataframe."""
    c1, c2 = st.columns([2, 3])
    with c1:
        search = st.text_input("🔍 Search (TCN / SO / Invoice ...)", key=f"{key_prefix}_search",
                               placeholder="Type to filter rows...")
    with c2:
        opts = df['Result'].unique().tolist()
        chosen = st.multiselect("Filter by status", opts, default=opts, key=f"{key_prefix}_filter")

    view = df[df['Result'].isin(chosen)]
    if search.strip():
        mask = view.apply(lambda row: row.astype(str).str.contains(search.strip(), case=False).any(), axis=1)
        view = view[mask]

    # Color-code Result column via pandas Styler
    def color_result(val):
        if '✅' in str(val):   return 'background-color:#F0FDF4; color:#16A34A; font-weight:600'
        if '⚠️' in str(val):  return 'background-color:#FFFBEB; color:#D97706; font-weight:600'
        if '🚫' in str(val):  return 'background-color:#FFF1F2; color:#E11D48; font-weight:600'
        if '❌' in str(val):  return 'background-color:#F8FAFC; color:#475569; font-weight:600'
        return ''

    styled = view.style.map(color_result, subset=['Result'])
    st.dataframe(styled, use_container_width=True, hide_index=True)
    st.caption(f"Showing {len(view)} of {len(df)} rows")
    return view


# ─── PAGE ──────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="header-banner">
    <div class="header-icon">🚛</div>
    <div>
        <div class="header-title">Trivedi ↔ Reliance Reconciliation</div>
        <div class="header-sub">Upload files below → Run Reconcile → Review results per tab</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── File uploaders ──
st.markdown('<div class="section-heading">Upload Files</div>', unsafe_allow_html=True)
c1, c2, c3, c4 = st.columns(4)

with c1:
    st.markdown('<div class="upload-label">Billing Portal<span class="upload-badge">XLSX · Required</span></div>', unsafe_allow_html=True)
    bill_file = st.file_uploader("", type=['xlsx'], key='bill', label_visibility='collapsed')

with c2:
    st.markdown('<div class="upload-label">Reliance Portal<span class="upload-badge">XLSX · Required</span></div>', unsafe_allow_html=True)
    rel_file = st.file_uploader("", type=['xlsx'], key='rel', label_visibility='collapsed')

with c3:
    st.markdown('<div class="upload-label">Vendor Summary<span class="upload-badge">CSV · Optional</span></div>', unsafe_allow_html=True)
    vendor_file = st.file_uploader("", type=['csv'], key='vendor', label_visibility='collapsed')

with c4:
    st.markdown('<div class="upload-label">Detention Sheet<span class="upload-badge">XLSX · Optional</span></div>', unsafe_allow_html=True)
    detention_file = st.file_uploader("", type=['xlsx'], key='detention', label_visibility='collapsed')

st.write("")

if rel_file and bill_file:
    if st.button("🔍  Run Reconciliation", type='primary'):
        with st.spinner("Processing files..."):
            try:
                bill_rows    = load_billing(bill_file.read())
                rel_rows     = load_reliance(rel_file.read())
                vendor_rows  = load_vendor_summary(vendor_file.read()) if vendor_file else None
                det_rows     = load_detention(detention_file.read()) if detention_file else None

                msg = f"✅  Loaded — {len(bill_rows)} Billing · {len(rel_rows)} Reliance"
                if vendor_rows: msg += f" · {len(vendor_rows)} Vendor"
                if det_rows:    msg += f" · {len(det_rows)} Detention rows"
                st.success(msg)

                tab1, tab2 = st.tabs(["📊  Reliance Reconciliation", "🚛  Detention LR Check"])

                # ── Tab 1 ──────────────────────────────────────────────────────
                with tab1:
                    results = reconcile(rel_rows, bill_rows, vendor_rows)
                    df = pd.DataFrame(results)
                    counts = df['Result'].value_counts().to_dict()

                    st.markdown('<div class="metric-row">', unsafe_allow_html=True)
                    mc1,mc2,mc3,mc4 = st.columns(4)
                    with mc1: metric_card(counts.get('✅ MATCHED',0),      "Matched",         "green")
                    with mc2: metric_card(counts.get('⚠️ MISMATCH',0),    "Mismatch",        "amber")
                    with mc3: metric_card(counts.get('🚫 REJECTED',0),    "Rejected",        "red")
                    with mc4: metric_card(counts.get('❌ NOT IN RELIANCE',0),"Not in Reliance","slate")
                    st.markdown('</div>', unsafe_allow_html=True)

                    if vendor_rows:
                        fb = df['Fallback Used'].astype(bool).sum()
                        st.info(f"🔁 Vendor Summary fallback used in **{fb}** rows")

                    st.markdown('<div class="section-heading">Results</div>', unsafe_allow_html=True)
                    view = render_table(df, 'tab1')
                    st.download_button("⬇️  Download CSV", df.to_csv(index=False).encode(),
                                       "reconciliation_result.csv", "text/csv")

                # ── Tab 2 ──────────────────────────────────────────────────────
                with tab2:
                    if det_rows is None:
                        st.markdown("""
                        <div class="empty-state">
                            <div class="empty-state-icon">📂</div>
                            <div class="empty-state-text">Detention Sheet nahi mila</div>
                            <div class="empty-state-sub">4th uploader me Detention XLSX upload karo aur dobara Reconcile karo</div>
                        </div>""", unsafe_allow_html=True)
                    else:
                        det_results = reconcile_detention(bill_rows, det_rows)
                        ddf = pd.DataFrame(det_results)
                        dcounts = ddf['Result'].value_counts().to_dict()

                        dc1,dc2,dc3 = st.columns(3)
                        with dc1: metric_card(dcounts.get('✅ MATCHED',0),           "LR Matched",   "green")
                        with dc2: metric_card(dcounts.get('⚠️ MISMATCH',0),         "Amount Diff",  "amber")
                        with dc3: metric_card(dcounts.get('❌ LR NOT IN DETENTION',0),"LR Not Found", "slate")

                        st.markdown('<div class="section-heading">Detention Results</div>', unsafe_allow_html=True)
                        render_table(ddf, 'tab2')
                        st.download_button("⬇️  Download Detention CSV",
                                           ddf.to_csv(index=False).encode(),
                                           "detention_result.csv", "text/csv")

            except Exception as e:
                st.error(f"Error: {e}")
                st.exception(e)
else:
    st.markdown("""
    <div class="empty-state">
        <div class="empty-state-icon">📋</div>
        <div class="empty-state-text">Billing + Reliance files upload karo to start</div>
        <div class="empty-state-sub">Vendor Summary aur Detention Sheet optional hain</div>
    </div>""", unsafe_allow_html=True)